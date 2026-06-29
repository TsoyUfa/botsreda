#!/usr/bin/env python3
"""Build 4-scenario purchase calculator for reference lot."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/anton_tsoy/Desktop/Обсидиан/9. Застройщики/Способы покупок/Калькулятор 4 сценария.xlsx"

# Reference lot (etalon)
PRICE = 6_000_000
AREA = 55
PV_PCT = 0.201
TERM_MONTHS = 360
BUILD_MONTHS = 24  # until final tranche / handover


def annuity_payment(principal: float, annual_rate: float, months: int) -> float:
    if principal <= 0:
        return 0.0
    if annual_rate <= 0:
        return principal / months
    r = annual_rate / 12
    return principal * r * (1 + r) ** months / ((1 + r) ** months - 1)


def total_paid_annuity(principal: float, annual_rate: float, months: int) -> float:
    return annuity_payment(principal, annual_rate, months) * months


def transh_scenario():
    """Transh: 100k first tranche, rest after 12 months, rate 21.7%, no markup."""
    pv = PRICE * PV_PCT
    credit = PRICE - pv
    tranche1 = 100_000
    tranche2 = credit - tranche1
    rate = 0.217

    pay_build = annuity_payment(tranche1, rate, TERM_MONTHS)
    pay_after = annuity_payment(credit, rate, TERM_MONTHS)

    # Interest during build (12 months on tranche1 only)
    interest_build = 0.0
    balance = tranche1
    r = rate / 12
    for _ in range(12):
        interest_build += balance * r
        balance -= annuity_payment(tranche1, rate, TERM_MONTHS) - balance * r

    total_interest = total_paid_annuity(credit, rate, TERM_MONTHS) - credit
    return {
        "name": "Траншевая ипотека (Сбер)",
        "price": PRICE,
        "markup": 0,
        "pv": pv,
        "credit": credit,
        "rate_note": "21,7% на весь срок",
        "pay_month_1": pay_build,
        "pay_year_1_avg": pay_build,
        "pay_year_3": pay_after,
        "pay_steady": pay_after,
        "total_interest": total_interest,
        "total_cost": PRICE + total_interest,
        "notes": (
            f"1-й транш 100 000 ₽, остаток через 12 мес. "
            f"Платёж до 2-го транша ~{pay_build:,.0f} ₽/мес, после ~{pay_after:,.0f} ₽/мес. "
            "Без удорожания. Арбитраж: ПВ−транш можно держать на депозите 18–20%."
        ),
    }


def subsidized_scenario():
    """Subsidized: 12.49% whole term, no markup (Bionika/Kondi Nova benchmark)."""
    pv = PRICE * PV_PCT
    credit = PRICE - pv
    rate = 0.1249
    pay = annuity_payment(credit, rate, TERM_MONTHS)
    total_interest = total_paid_annuity(credit, rate, TERM_MONTHS) - credit
    return {
        "name": "Субсидия 12,49% (весь срок, без удорожания)",
        "price": PRICE,
        "markup": 0,
        "pv": pv,
        "credit": credit,
        "rate_note": "12,49% на весь срок",
        "pay_month_1": pay,
        "pay_year_1_avg": pay,
        "pay_year_3": pay,
        "pay_steady": pay,
        "total_interest": total_interest,
        "total_cost": PRICE + total_interest,
        "notes": "Бенчмарк: Бионика Парк, Конди Нова, Совкомбанк 12,49%. Предсказуемая переплата.",
    }


def sovcom_scenario():
    """Sovcom reduced payment: 4.4% year 1, then 19.99%."""
    pv = PRICE * 0.2001
    credit = PRICE - pv
    rate1 = 0.044
    rate2 = 0.1999
    pay1 = annuity_payment(credit, rate1, TERM_MONTHS)
    pay2 = annuity_payment(credit, rate2, TERM_MONTHS)

    # Approximate total interest: 12 months at rate1 schedule + rest at rate2
    # Simplified: recalculate remaining balance after 12 payments at rate1, then rate2
    r1 = rate1 / 12
    balance = credit
    interest_total = 0.0
    pmt1 = pay1
    for _ in range(12):
        interest = balance * r1
        principal_part = pmt1 - interest
        interest_total += interest
        balance -= principal_part

    r2 = rate2 / 12
    remaining_months = TERM_MONTHS - 12
    pmt2 = annuity_payment(balance, rate2, remaining_months)
    for _ in range(remaining_months):
        interest = balance * r2
        principal_part = pmt2 - interest
        interest_total += interest
        balance -= principal_part

    return {
        "name": "Сниженный платёж Совкомбанка",
        "price": PRICE,
        "markup": 0,
        "pv": pv,
        "credit": credit,
        "rate_note": "4,4% → 19,99% (12 мес. льгота)",
        "pay_month_1": pay1,
        "pay_year_1_avg": pay1,
        "pay_year_3": pay2,
        "pay_steady": pay2,
        "total_interest": interest_total,
        "total_cost": PRICE + interest_total,
        "notes": (
            f"Платёж год 1 ~{pay1:,.0f} ₽, с 13-го мес. ~{pay2:,.0f} ₽. "
            "Маткапитал в ПВ запрещён. Жёсткий «обрыв» платежа."
        ),
    }


def installment_scenario():
    """Installment: Arkhstroy PV 30%, +10k/m2 markup, 40k/month, rest to mortgage."""
    markup_per_sqm = 10_000
    markup = markup_per_sqm * AREA
    price_total = PRICE + markup
    pv = price_total * 0.30
    remainder = price_total - pv
    monthly = 40_000
    months_pay = 9  # until 06.2027 approx from Jun 2026
    paid_monthly = monthly * months_pay
    rest_for_mortgage = remainder - paid_monthly

    # Assume rest goes to mortgage at 12.49% after installment period
    rate = 0.1249
    mortgage_months = TERM_MONTHS - months_pay
    pay_mortgage = annuity_payment(rest_for_mortgage, rate, mortgage_months)
    mortgage_interest = total_paid_annuity(rest_for_mortgage, rate, mortgage_months) - rest_for_mortgage

    return {
        "name": "Рассрочка Архстрой (ПВ 30%, +10к/м²)",
        "price": price_total,
        "markup": markup,
        "pv": pv,
        "credit": rest_for_mortgage,
        "rate_note": f"Рассрочка 0%, остаток → ипотека 12,49%",
        "pay_month_1": monthly,
        "pay_year_1_avg": monthly,
        "pay_year_3": pay_mortgage,
        "pay_steady": pay_mortgage,
        "total_interest": mortgage_interest,
        "total_cost": price_total + mortgage_interest,
        "notes": (
            f"Удорожание +{markup:,.0f} ₽ ({markup_per_sqm:,} ₽/м² × {AREA} м²). "
            f"9 мес. по {monthly:,} ₽, остаток {rest_for_mortgage:,.0f} ₽ → ипотека. "
            "Ключи после 100% оплаты."
        ),
    }


def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2F5496")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_money(cell):
    cell.number_format = '#,##0 ₽'


def style_pct(cell):
    cell.number_format = '0.0%'


def build():
    scenarios = [
        transh_scenario(),
        subsidized_scenario(),
        sovcom_scenario(),
        installment_scenario(),
    ]

    wb = Workbook()

    # --- Sheet 1: Inputs ---
    ws_in = wb.active
    ws_in.title = "Параметры лота"
    ws_in["A1"] = "Эталонный лот для сравнения"
    ws_in["A1"].font = Font(bold=True, size=14)
    inputs = [
        ("Цена квартиры (базовая)", PRICE),
        ("Площадь, м²", AREA),
        ("ПВ, %", PV_PCT),
        ("Срок кредита, лет", TERM_MONTHS / 12),
        ("Срок стройки до 2-го транша, мес.", 12),
        ("Дата расчёта", "22.06.2026"),
        ("Источник", "База «Способы покупок», Уфа"),
    ]
    for i, (label, val) in enumerate(inputs, start=3):
        ws_in[f"A{i}"] = label
        ws_in[f"B{i}"] = val
        ws_in[f"A{i}"].font = Font(bold=True)
    ws_in.column_dimensions["A"].width = 38
    ws_in.column_dimensions["B"].width = 22

    # --- Sheet 2: Comparison ---
    ws = wb.create_sheet("Сравнение 4 сценариев")
    headers = [
        "Сценарий",
        "Цена сделки",
        "Удорожание",
        "ПВ",
        "Сумма кредита / остаток",
        "Ставка / условие",
        "Платёж, мес. 1",
        "Средний платёж, год 1",
        "Платёж, год 3",
        "Платёж после стабилизации",
        "Переплата по %",
        "Полная стоимость",
        "Комментарий",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        style_header(c)

    for row_idx, s in enumerate(scenarios, start=2):
        row = [
            s["name"],
            s["price"],
            s["markup"],
            s["pv"],
            s["credit"],
            s["rate_note"],
            s["pay_month_1"],
            s["pay_year_1_avg"],
            s["pay_year_3"],
            s["pay_steady"],
            s["total_interest"],
            s["total_cost"],
            s["notes"],
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col in (2, 3, 4, 5, 7, 8, 9, 10, 11, 12):
                style_money(cell)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    widths = [32, 14, 14, 14, 16, 22, 14, 14, 14, 16, 14, 14, 48]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet 3: Summary insights ---
    ws2 = wb.create_sheet("Выводы")
    ws2["A1"] = "Краткие выводы по эталонному лоту"
    ws2["A1"].font = Font(bold=True, size=14)

    best_year1 = min(scenarios, key=lambda x: x["pay_month_1"])
    best_total = min(scenarios, key=lambda x: x["total_cost"])
    worst_total = max(scenarios, key=lambda x: x["total_cost"])

    insights = [
        "",
        f"Самый низкий платёж в год 1: {best_year1['name']} — {best_year1['pay_month_1']:,.0f} ₽/мес.",
        f"Минимальная полная стоимость: {best_total['name']} — {best_total['total_cost']:,.0f} ₽",
        f"Максимальная полная стоимость: {worst_total['name']} — {worst_total['total_cost']:,.0f} ₽",
        "",
        "Траншевая: лучший кэшфлоу на стройке, но рыночная ставка 21,7% — дорого за весь срок.",
        "Субсидия 12,49%: оптимальна по переплате при отсутствии льготных программ.",
        "Совкомбанк: низкий вход, но обрыв платежа на 2-й год.",
        "Рассрочка с удорожанием: скрытая переплата в цене + ипотека на остаток.",
        "",
        "При наличии семейной ипотеки (3,5–6%) все 4 сценария вторичны.",
    ]
    for i, line in enumerate(insights, start=3):
        ws2[f"A{i}"] = line
    ws2.column_dimensions["A"].width = 90

    # --- Sheet 4: Formulas reference ---
    ws3 = wb.create_sheet("Формулы")
    ws3["A1"] = "Справка по расчёту"
    ws3["A1"].font = Font(bold=True, size=14)
    formulas = [
        ("Аннуитетный платёж", "П = S × r × (1+r)^n / ((1+r)^n − 1), r = годовая/12"),
        ("Траншевая", "Проценты только на выданный транш; 2-й транш через 12 мес."),
        ("Субсидия", "Ставка 12,49% на весь срок, цена без удорожания"),
        ("Совкомбанк", "12 мес. по 4,4%, далее пересчёт на 19,99% на остаток срока"),
        ("Рассрочка", "Цена + 10 000₽/м²; 9×40 000₽; остаток → ипотека 12,49%"),
    ]
    for i, (name, desc) in enumerate(formulas, start=3):
        ws3[f"A{i}"] = name
        ws3[f"B{i}"] = desc
        ws3[f"A{i}"].font = Font(bold=True)
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 60

    wb.save(OUT)
    print(f"Saved: {OUT}")
    for s in scenarios:
        print(
            f"{s['name']}: год1={s['pay_year_1_avg']:,.0f}, "
            f"стаб={s['pay_steady']:,.0f}, полная={s['total_cost']:,.0f}"
        )


if __name__ == "__main__":
    build()
